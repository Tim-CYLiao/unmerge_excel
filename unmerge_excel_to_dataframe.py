#!/usr/bin/env python
#-*- coding -*-
#Reference is from https://github.com/zanran/unMergeExcelCell.


import os
import sys
import openpyxl
import pandas as pd


__doc__ = '''
Function:
    unmerge_excel_to_df_dict(path): For unmerge excel into dictionary of dataframe.
    unmerge_ws_to_df(book,sheetnames): Unmerge specific worksheet of openpyxl's workbook object into a dataframe
Unmerged cells are filled by value of original merged cells.

'''



def usage():
    print(__doc__)
    exit(-1)
    
#unmerge_excel_to_df_dict(path): For unmerge excel into dict of dataframe.
def unmerge_excel_to_df_dict(path):
    
    if not os.path.exists(path):
        print(("Could not find the excel file: " % path))
        return
    
    book = openpyxl.load_workbook(path,read_only = False)
    df_dict={}
    
    for sheet_name in book.sheetnames:
        df_dict.update( {sheet_name :unmerge_ws_to_df( book, sheet_name) } )
    print(df_dict)
    return df_dict


#unmerge_ws_to_df(book,sheetnames): Unmerge specific worksheet of openpyxl's workbook object into a dataframe
def unmerge_ws_to_df(book,sheetnames):

    rd_sheet = book[sheetnames]
    # over write for merged cells
    data_df = pd.DataFrame(rd_sheet.values)
    
    for merged_cell in rd_sheet.merged_cells.ranges:
        merged_cell_coord = str(merged_cell)        
        value_coord = str(merged_cell_coord.split(":")[0])# get the index of first cell
        cell_value = rd_sheet[value_coord].value
        min_col_t, min_row_t, max_col_t, max_row_t = merged_cell#each data type is tuple ex (min_col,A1:A2)
        min_col, min_row, max_col, max_row = min_col_t[1], min_row_t[1], max_col_t[1], max_row_t[1]
        #range(x,y)generate sequence of x:(y-1),so add them back
        
        for rowx in range(min_row, max_row+1):
            
            for colx in range(min_col, max_col+1):
                data_df.loc[rowx-1,colx-1] = cell_value       
    return data_df


if __name__ == "__main__":
    current_dir = os.getcwd()
    path = os.path.join(current_dir, 'test_merge-file.xlsx')
    unmerge_excel_to_df_dict(path)
