#!/usr/bin/env python
#-*- coding -*-
#Reference is from https://github.com/zanran/unMergeExcelCell.


import os
import sys
import openpyxl
import pandas as pd


__doc__ = '''
Function:
    unmerge_excel(path): For unmerge excel and  save it.
    now some of merged cells couldn't be unmerged,need to fix it
Unmerged cells are filled by value of original merged cells.

'''



def usage():
    print(__doc__)
    exit(-1)
    
#unmerge_excel(path): For unmerge excel and save it.
def unmerge_excel(path):
    
    if not os.path.exists(path):
        print(("Could not find the excel file: " % path))
        return
    
    book = openpyxl.load_workbook(path,read_only = False)

    for rd_sheet_name in book.sheetnames:
        rd_sheet = book.active
        rd_sheet = book[rd_sheet_name]
        # over write for merged cells
        rd_sheet.merged_cells.ranges
        for merged_cell in rd_sheet.merged_cells.ranges:
            rd_sheet.unmerge_cells(merged_cell.coord)

            min_col_t, min_row_t, max_col_t, max_row_t = merged_cell#each data type is tuple ex (min_col,A1:A2)
            min_col, min_row, max_col, max_row = min_col_t[1], min_row_t[1], max_col_t[1], max_row_t[1]
            cell_value = rd_sheet.cell(row = min_row, column = min_col).value
            #range(x,y)generate sequence of x:(y-1),so add them back
            for rowx in range(min_row, max_row+1):
                for colx in range(min_col, max_col+1):
                    rd_sheet.cell(row = rowx, column = colx, value = cell_value)
                    assert rd_sheet.cell(row = rowx, column = colx).value == cell_value, "Error at %rowx, %colx with %cell_value"
    (origin_file, ext) = os.path.splitext(path)

    unmerge_excel_file = "unmerged.xlsx"
    book.save(unmerge_excel_file)


if __name__ == "__main__":
    current_dir = os.getcwd()
    path = os.path.join(current_dir, 'test_merge-file.xlsx')
    unmerge_excel(path)
